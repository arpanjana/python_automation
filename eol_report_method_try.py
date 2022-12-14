""""
1. to open the excel sheet and be able to extract the data from the column
2. create a loop for the
to read the excel sheet
pick those value and feed in respective cycle
find the total number of li

"""
from email import header
import sys
import xlsxwriter
import time
import os
import xlrd 
import openpyxl as xl


# the location of the file 
loc = r"C:\Users\ajana\Desktop\python-project\OUTPUT_16_08_2022_eol_setup_asc_logs1_1.xlsx"
loc1 = r"C:\Users\ajana\Desktop\python-project\testing_scripts"
excel_sheet_list = os.listdir(loc1)
"""
11_08_2022_NHW_23.40

"""
date = []
build_hardware = []
build_no = []

for f in excel_sheet_list:
    date.append(f[0:10])
    build_hardware.append(f[11:14])
    build_no.append(f[15:20])



OUTPUT="OUTPUT_consolidated_report"+".xlsx"
workbook = xlsxwriter.Workbook(OUTPUT)

dest_path = os.path.abspath(OUTPUT)

worksheet_version_history = workbook.add_worksheet("Version History")
worksheet_eol_summary = workbook.add_worksheet("EOL summary")

# setting up the version history for the excel files
border_format = workbook.add_format({
    'border': 2,
    'align':'left',
    'font_size': 15
})
worksheet_version_history.write('C2','Project Name',border_format)
worksheet_version_history.write('C3','Document Title',border_format)
worksheet_version_history.write('C4','Date',border_format)
worksheet_version_history.write('C5','Version',border_format)
worksheet_version_history.write('C6','Status',border_format)
worksheet_version_history.write('C7','Owner',border_format)
worksheet_version_history.write('C8','Approved',border_format)


worksheet_version_history.write('B12','Version',border_format)
worksheet_version_history.write('C12','Date',border_format)
worksheet_version_history.write('D12','Change from Previous',border_format)


cnt_version_row = 13

for i in range(len(excel_sheet_list)):
    worksheet_version_history.write(cnt_version_row,1,(i+1)/10,border_format)
    worksheet_version_history.write(cnt_version_row,2,date[i],border_format)
    title = "EOL test report on " + build_hardware[i] + build_no[i]
    worksheet_version_history.write(cnt_version_row,3,title,border_format)

# setting up the eol sumary for the excel files

header_format=workbook.add_format({'top':1,'right':1,'left':1,'bottom':1,'align':'center'})
header_format.set_bg_color("yellow")

cnt_summary_row = 3
for i in range(len(excel_sheet_list)):
    lr = 'B'+str(cnt_summary_row)
    hr = 'C'+str(cnt_summary_row)
    tr = lr+":"+hr
    worksheet_eol_summary.merge_range(tr,'PSA CARUSO Test Case Execution Summary',header_format)
    cnt_summary_row += 1
    worksheet_eol_summary.write(cnt_summary_row,1,'EOL Tool Version',header_format)
    worksheet_eol_summary.write(cnt_summary_row,2,'0.2',header_format)
    cnt_summary_row +=1
    worksheet_eol_summary.write(cnt_summary_row,1,'Brand',header_format)
    worksheet_eol_summary.write(cnt_summary_row,2,build_hardware[i],header_format)
    cnt_summary_row += 1
    worksheet_eol_summary.write(cnt_summary_row,1,'Build Version',header_format)
    worksheet_eol_summary.write(cnt_summary_row,2,build_no[i],header_format)
    cnt_summary_row +=1
    worksheet_eol_summary.write(cnt_summary_row,1,'Number of cycles',header_format)
    cnt_summary_row += 5



workbook.close()







# for putting the timings for 


for files in excel_sheet_list:
    # new_loc = os.path.abspath(files)
    new_loc = os.path.join(loc1,files)
    wb1 = xl.load_workbook(new_loc)
    ws1 = wb1.worksheets[0]

    wb2 = xl.load_workbook(dest_path)
    ws2 = wb2.active

    mr = ws1.max_row
    mc = ws1.max_column

    #copying the cell values from source 
    #excel file to destination excel file
    for i in range(1,mr+1):
        for j in range(1,mc+1):
            # reading cell value from source excel file
            c = ws1.cell(row = i,column = j)

            # writing the read value to destination excel file
            ws2.cell(row = i,column = j).value = c.value

    # saving the destination excel file
    wb2.save(str(files))




